# coding=utf-8
from __future__ import print_function
from __future__ import print_function
from __future__ import print_function
import os
import sys
import csv
import networkx as nx
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect, HttpResponse
from networkx import NetworkXError
from django.shortcuts import render_to_response, render
from .models import NetworkElement, FiberRelationship, ConvergeNE, Result, DetailResult, LongSingleTable, NeTable, \
    CneTable, RingTable
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')


def index(request):
    return render(request, 'analyze/index.html')


def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user:
            if user.is_active:
                login(request, user)
                return HttpResponseRedirect('/analyze/')
            else:
                return HttpResponse("Your account is disabled.")
        else:
            print("Invalid login details: {0}, {1}".format(username, password))
            return HttpResponse("Invalid login details supplied.")
    else:
        return render(request, 'analyze/login.html', {})


def readme(request):
    return render_to_response('help.html')


@login_required
def user_logout(request):
    logout(request)
    return HttpResponseRedirect('/analyze/')


@login_required
def upload(request):
    if os.path.exists('E:\\upload\\neinfo.csv'):
        os.remove('E:\\upload\\neinfo.csv')
    if os.path.exists('E:\\upload\\cneinfo.csv'):
        os.remove('E:\\upload\\cneinfo.csv')
    if os.path.exists('E:\\upload\\ne2ne.csv'):
        os.remove('E:\\upload\\ne2ne.csv')
    if request.method == "POST":
        neinfo = request.FILES.get("neinfo", None)  # 获取上传的文件，如果没有文件，则默认为None
        ne2ne = request.FILES.get("ne2ne", None)  # 获取上传的文件，如果没有文件，则默认为None
        cneinfo = request.FILES.get("cneinfo", None)  # 获取上传的文件，如果没有文件，则默认为None
        if not neinfo and ne2ne and cneinfo:
            return HttpResponse("no files for upload!")
        destination_neinfo = open(os.path.join("E:\\upload", neinfo.name), 'wb+')  # 打开特定的文件进行二进制的写操作
        destination_ne2ne = open(os.path.join("E:\\upload", ne2ne.name), 'wb+')  # 打开特定的文件进行二进制的写操作
        destination_cneinfo = open(os.path.join("E:\\upload", cneinfo.name), 'wb+')  # 打开特定的文件进行二进制的写操作
        for chunk in neinfo.chunks():  # 分块写入文件
            destination_neinfo.write(chunk)
        destination_neinfo.close()
        for chunk in ne2ne.chunks():  # 分块写入文件
            destination_ne2ne.write(chunk)
        destination_ne2ne.close()
        for chunk in cneinfo.chunks():  # 分块写入文件
            destination_cneinfo.write(chunk)
        destination_cneinfo.close()
        print(u"成功上传数据源！")
        return render_to_response('success.html')


@login_required
def produce_result(request):
    NetworkElement.objects.all().delete()
    FiberRelationship.objects.all().delete()
    ConvergeNE.objects.all().delete()
    Result.objects.all().delete()
    DetailResult.objects.all().delete()
    FiberRelationship.objects.all().delete()
    LongSingleTable.objects.all().delete()
    NeTable.objects.all().delete()
    CneTable.objects.all().delete()
    RingTable.objects.all().delete()
    print(u"清理数据完成！")
    # 导入网元信息表中的普通网元到数据库
    with open("E:\\upload\\neinfo.csv", "r") as neinfo_csvfile:
        neinfo_read = csv.reader(neinfo_csvfile)
        access_ne_type = ('OptiX PTN 910', 'OptiX PTN 950', 'OptiX PTN 960', 'OptiX PTN 1900')
        neinfo_list = []
        nelist = [ne_row for ne_row in neinfo_read]
        if nelist[7][0] == u'网元名称' and nelist[7][1] == u'网元类型' and nelist[7][9] == u'所属子网':
            for nelist_range in range(8, len(nelist)):
                if nelist[nelist_range][1] in access_ne_type:
                    neinfo_list.append(NetworkElement(ne_name=nelist[nelist_range][0], ne_type=nelist[nelist_range][1],
                                                      ring_name=nelist[nelist_range][9],
                                                      ring_region=(nelist[nelist_range][9].decode('utf-8'))[0:2]))
            NetworkElement.objects.bulk_create(neinfo_list)
            print(u"成功导入网元信息表！")
        else:
            print(u'网元信息表读取有误！')

    # 导入汇聚网元信息表到数据库
    with open("E:\\upload\\cneinfo.csv", "r") as cneinfo_csvfile:
        cneinfo_read = csv.reader(cneinfo_csvfile)
        cneinfo_list = []
        cnelist = [cne_row for cne_row in cneinfo_read]
        if cnelist[0][1] == u'网元名称' and cnelist[0][2] == u'所属子网':
            for cnelist_range in range(1, len(cnelist)):
                cneinfo_list.append(ConvergeNE(cne_name=cnelist[cnelist_range][1], cne_type='390000',
                                               ring_name=cnelist[cnelist_range][2],
                                               ring_region=cnelist[cnelist_range][0]))
            ConvergeNE.objects.bulk_create(cneinfo_list)
            print(u"成功导入汇聚网元信息表！")
        else:
            print(u'网元信息表读取有误！')

    # 导入纤缆连接关系表到数据库
    with open("E:\\upload\\ne2ne.csv", "r") as ne2ne_csvfile:
        ne2ne_read = csv.reader(ne2ne_csvfile)
        ne2ne_list = []
        ne2nelist = [ne2ne_row for ne2ne_row in ne2ne_read]
        if ne2nelist[7][5] == u'源网元' and ne2nelist[7][7] == u'宿网元' and ne2nelist[7][16] == u'备注':
            for ne2nelist_range in range(8, len(ne2nelist)):
                if ne2nelist[ne2nelist_range][16] == '1':
                    ne2ne_list.append(
                        FiberRelationship(source=ne2nelist[ne2nelist_range][5], target=ne2nelist[ne2nelist_range][7],
                                          edge_weight=1))
                elif ne2nelist[ne2nelist_range][16] != '1' and ne2nelist[ne2nelist_range][16] != '0':
                    ne2ne_list.append(
                        FiberRelationship(source=ne2nelist[ne2nelist_range][5], target=ne2nelist[ne2nelist_range][7],
                                          edge_weight=1000))
            FiberRelationship.objects.bulk_create(ne2ne_list)
            print(u"成功导入纤缆连接关系表！")
        else:
            print(u'网元信息表读取有误！')

    # 以下为计算参数配置-----------------------------
    # 超大接入环接入网元阈值
    big_access_num_config = 12
    # 长单链网元阈值
    long_single_chain_num_config = 4
    # 超大汇聚节点阈值
    big_converge_node_point_config = 0
    # 以上为计算参数配置-----------------------------

    # 在汇聚网元表中筛选出所有环名称
    ring_name_set = ConvergeNE.objects.values_list('ring_name', flat=True).order_by('ring_name').distinct()
    # 所有接入环接入网元数
    total_access_ne_num = []
    # 所有接入环成环接入网元数
    total_ring_access_ne_num = []
    # 所有双归网元表
    total_double_ne_num = []
    # 遍历ring_name_set，到NetworkElement和ConvergeNE两张表中搜索ring_name_set[]得到某一环下所有网元
    for rns in ring_name_set:
        # 获取某环接入环接入网元列表
        access_ne_list = list(NetworkElement.objects.filter(ring_name=rns).values_list('ne_name', flat=True))
        # 获取某环接入环汇聚网元列表
        converge_ne_list = list(ConvergeNE.objects.filter(ring_name=rns).values_list('cne_name', flat=True))
        # 将该环接入网元数量添加至总数量列表
        total_access_ne_num.append(len(access_ne_list))
        # 获取某环接入环全部网元列表
        ne_list = access_ne_list + converge_ne_list
        # 获取源跟宿均在该环全部网元列表中的纤缆连接关系列表
        fiber_relationship_list = FiberRelationship.objects.filter(source__in=ne_list).filter(target__in=ne_list) \
            .values_list('source', 'target')
        # 实例化一个空图
        g = nx.Graph()
        # 定义用来存放成环网元的列表
        path_list = []
        # 将路径列表加入到图中
        g.add_edges_from(fiber_relationship_list)
        # 成环汇聚网元
        ring_cne = []
        # --------------------计算成环率和超大接入环-------------------
        # 找出接入环中任意两个汇聚网元之间的所有路径
        for a in range(0, len(converge_ne_list) - 1):
            for b in range(a + 1, len(converge_ne_list)):
                source_ne = converge_ne_list[a]
                target_ne = converge_ne_list[b]
                try:
                    # 判断路径非空
                    if nx.all_simple_paths(g, source=source_ne, target=target_ne):
                        if source_ne not in ring_cne:
                            ring_cne.append(source_ne)
                        elif target_ne not in ring_cne:
                            ring_cne.append(target_ne)
                        paths = nx.all_simple_paths(g, source=source_ne, target=target_ne)
                        for pth in paths:
                            # 去除路径中包含汇聚的情况
                            if len(set(pth) & set(converge_ne_list)) == 2:
                                if len(pth) > 2:
                                    pth_str = '，'.join(pth)
                                    DetailResult.objects.create(ring_name=rns, arp=pth_str)
                                    # 判断超大接入环并写入数据库BAR数据表中
                                    if (len(pth) - 2) >= big_access_num_config:
                                        DetailResult.objects.create(ring_name=rns, ne_num=len(pth) - 2, bar_ne=pth_str)
                                    # 将该路径中的网元加入到该环成环网元列表中，剔除首尾的汇聚网元
                                    for c in range(1, len(pth) - 1):
                                        if pth[c] not in path_list:
                                            path_list.append(pth[c])
                except NetworkXError as nxe:
                    print(nxe.message)
                    DetailResult.objects.get_or_create(ring_name=rns, msg=nxe.message)
                    continue
        # 获取每一个环成环网元列表
        ring_ne_list = path_list
        ring_str = '，'.join(ring_ne_list)
        # 将该环的成环率写入数据库ARR数据表中
        try:
            DetailResult.objects.create(ring_name=rns, arr=float(len(ring_ne_list)) / float(len(access_ne_list)),
                                        arr_ne=ring_str)
        except ZeroDivisionError:
            print('ZeroDivisionError')
            DetailResult.objects.get_or_create(ring_name=rns, msg='ZeroDivisionError')
            continue
        # 将该环成环网元的数量添加到总列表中
        total_ring_access_ne_num.append(len(ring_ne_list))
        # 获取每一个环未成环网元列表
        no_ring_ne_list = list(set(access_ne_list) - set(ring_ne_list))

        # --------------------计算长单链-------------------
        # 成环接入汇聚网元列表
        new_ring = ring_cne + ring_ne_list
        # 计算长单链
        for rnl in new_ring:
            for nrnl in no_ring_ne_list:
                try:
                    if nx.all_simple_paths(g, source=rnl, target=nrnl):
                        for nrnp in nx.all_simple_paths(g, source=rnl, target=nrnl):
                            if len(set(nrnp) & set(new_ring)) == 1 and len(
                                    nrnp) >= long_single_chain_num_config:
                                nrnp_str = '，'.join(nrnp)
                                DetailResult.objects.create(ring_name=rns, lsc_num=len(nrnp), lsc_ne=nrnp_str)
                except NetworkXError as nxe:
                    print(nxe.message)
                    DetailResult.objects.get_or_create(ring_name=rns, msg=nxe.message)
                    continue
        # 处理重复的长单链
        lsc_original = DetailResult.objects.filter(ring_name=rns).values_list('id', 'lsc_num', 'lsc_ne').order_by(
            '-lsc_num')
        # print('=======lsc_original=======')
        # print(lsc_original)
        for i in range(0, len(lsc_original) - 1):
            for j in range(i + 1, len(lsc_original)):
                big = lsc_original[i]
                lessbig = lsc_original[j]
                if big[2] and lessbig[2]:
                    big_ne_list = str(big[2]).split('，')
                    lessbig_ne_list = str(lessbig[2]).split('，')
                    if set(big_ne_list).issuperset(set(lessbig_ne_list)):
                        if DetailResult.objects.filter(id=lessbig[0]):
                            DetailResult.objects.filter(id=lessbig[0]).delete()

        # --------------------计算双归率-------------------
        # 定义单归网元列表
        single_accsess_ne = []
        # 计算单归
        for cnl in converge_ne_list:
            for nrnl in no_ring_ne_list:
                try:
                    if nx.all_simple_paths(g, source=cnl, target=nrnl):
                        for nrnp in nx.all_simple_paths(g, source=cnl, target=nrnl):
                            if len(set(nrnp) & set(ring_ne_list + converge_ne_list)) == 1:
                                for c in range(0, len(nrnp)):
                                    if nrnp[c] not in single_accsess_ne:
                                        single_accsess_ne.append(nrnp[c])
                except NetworkXError as nxe:
                    print(nxe.message)
                    DetailResult.objects.get_or_create(ring_name=rns, msg=nxe.message)
                    continue
        double_accsess_ne = list(set(access_ne_list) - (set(single_accsess_ne) - set(converge_ne_list)))
        double_str = '，'.join(double_accsess_ne)
        # 将该环双归率写入数据库DR数据表中
        try:
            DetailResult.objects.create(ring_name=rns, dr=float(len(double_accsess_ne)) / float(len(access_ne_list)),
                                        dr_ne=double_str)
        except ZeroDivisionError:
            print('ZeroDivisionError')
            DetailResult.objects.get_or_create(ring_name=rns, msg='ZeroDivisionError')
            continue

        total_double_ne_num.append(len(double_accsess_ne))

        # --------------------计算超大汇聚节点-------------------
        for a in range(0, len(converge_ne_list)):
            single_single_ne = []
            for b in range(0, len(no_ring_ne_list)):
                try:
                    if nx.all_simple_paths(g, source=converge_ne_list[a], target=no_ring_ne_list[b]):
                        for ph in nx.all_simple_paths(g, source=converge_ne_list[a], target=no_ring_ne_list[b]):
                            if len(set(ph) & set(ring_ne_list + converge_ne_list)) == 1:
                                for p in range(1, len(ph)):
                                    if ph[p] not in single_single_ne:
                                        single_single_ne.append(ph[p])
                except NetworkXError as nxe:
                    print(nxe.message)
                    DetailResult.objects.get_or_create(ring_name=rns, msg=nxe.message)
                    continue
            point = len(single_single_ne) + float(len(double_accsess_ne)) / float(len(converge_ne_list))
            if point >= big_converge_node_point_config:
                DetailResult.objects.create(ring_name=rns, cne_point=point, bcne_cne=converge_ne_list[a])

    ring_rate = float(sum(total_ring_access_ne_num)) / float(sum(total_access_ne_num))
    double_rate = float(sum(total_double_ne_num)) / float(sum(total_access_ne_num))
    Result.objects.create(total_arr=ring_rate, total_dr=double_rate)

    # [(id, u''),(id, u''),....]
    arplist = DetailResult.objects.values_list('id', 'arp')
    for al in arplist:
        weightlist = []
        allist = str(al[1]).split('，')
        for x in range(0, len(allist) - 1):
            if FiberRelationship.objects.filter(source=allist[x]).filter(target=allist[x + 1]):
                v = FiberRelationship.objects.filter(source=allist[x]).filter(target=allist[x + 1]).values_list(
                    'edge_weight', flat=True)
                weightlist.append(v)
            elif FiberRelationship.objects.filter(source=allist[x + 1]).filter(target=allist[x]):
                v = FiberRelationship.objects.filter(source=allist[x + 1]).filter(target=allist[x]).values_list(
                    'edge_weight', flat=True)
                weightlist.append(v)
        # print(weightlist)

        if len(weightlist) > 0 and DetailResult.objects.filter(id=al[0]):
            if str(weightlist[0]) == "[u'1000']" and DetailResult.objects.filter(id=al[0]):
                for wln in range(1, len(weightlist)):
                    if str(weightlist[wln]) != str(weightlist[0]) and wln < len(
                            weightlist) - 1 and DetailResult.objects.filter(id=al[0]):
                        for wlnl in range(wln + 1, len(weightlist)):
                            if str(weightlist[wlnl]) == str(weightlist[0]) and DetailResult.objects.filter(id=al[0]):
                                DetailResult.objects.filter(id=al[0]).delete()
            elif str(weightlist[0]) == "[u'1']" and DetailResult.objects.filter(id=al[0]):
                for wln in range(1, len(weightlist)):
                    if str(weightlist[wln]) != str(weightlist[0]) and wln < len(
                            weightlist) - 1 and DetailResult.objects.filter(id=al[0]):
                        for wln1 in range(wln + 1, len(weightlist)):
                            if str(weightlist[wln1]) != str(weightlist[wln]) and wln1 < len(
                                    weightlist) - 1 and DetailResult.objects.filter(id=al[0]):
                                for wln2 in range(wln1 + 1, len(weightlist)):
                                    if str(weightlist[wln2]) != str(weightlist[wln1]) and DetailResult.objects.filter(
                                            id=al[0]):
                                        DetailResult.objects.filter(id=al[0]).delete()
    print(u"成功生成调试表！")
    # 生成长单链表
    lstlist = []
    lst = DetailResult.objects.values_list('id', 'ring_name', 'lsc_num', 'lsc_ne')
    for l in lst:
        if len(l[2]) != 0:
            lsc_region = l[1][0:2]
            lstlist.append(LongSingleTable(region=lsc_region, longsinglepath=l[3], nbr=l[2]))
    LongSingleTable.objects.bulk_create(lstlist)
    print(u"成功生成长单链表！")

    # 生成汇聚表
    cnetablelist = []
    cneset = list(set(ConvergeNE.objects.values_list('cne_name', flat=True)))
    for cs in cneset:
        cspointlist = DetailResult.objects.filter(bcne_cne=cs).values_list('cne_point', flat=True)
        cspointlistnbr = [float(x) for x in cspointlist]
        cspoint = sum(cspointlistnbr)
        ct_region = ConvergeNE.objects.filter(cne_name=cs).values_list('ring_region', flat=True)[0]
        if cspoint >= 80:
            cnetablelist.append(CneTable(region=ct_region, cne_name=cs, cne_nenbr=cspoint, is_big_cne='是'))
        else:
            cnetablelist.append(CneTable(region=ct_region, cne_name=cs, cne_nenbr=cspoint, is_big_cne='否'))
    CneTable.objects.bulk_create(cnetablelist)
    print(u"成功生成汇聚网元表！")

    # 生成成环表
    rt_ring = set(DetailResult.objects.values_list('ring_name', flat=True))
    ringtablelist = []
    for rr in rt_ring:
        rt = DetailResult.objects.filter(ring_name=rr).exclude(arp='').values_list('arp', flat=True)
        slavenbr = len(rt) - 1
        for r in range(0, len(rt)):
            weightnewlist = []
            pointer = 0
            rlist = str(rt[r]).split('，')
            for p in range(0, len(rlist) - 1):
                if FiberRelationship.objects.filter(source=rlist[p]).filter(target=rlist[p + 1]):
                    q = FiberRelationship.objects.filter(source=rlist[p]).filter(target=rlist[p + 1]).values_list(
                        'edge_weight', flat=True)
                    weightnewlist.append(q)
                elif FiberRelationship.objects.filter(source=rlist[p + 1]).filter(target=rlist[p]):
                    q = FiberRelationship.objects.filter(source=rlist[p + 1]).filter(target=rlist[p]).values_list(
                        'edge_weight', flat=True)
                    weightnewlist.append(q)
            for wnl in weightnewlist:
                if str(wnl) == u"<QuerySet [u'1000']>":
                    pointer = 1
                    break
            if pointer == 0:
                if (len(rlist) - 2) >= 12:
                    ringtablelist.append(RingTable(region=rr[0:2], ring_name=rr, arp=rt[r], arp_nbr=len(rlist) - 2,
                                                   is_big_ring='是'))
                else:
                    ringtablelist.append(RingTable(region=rr[0:2], ring_name=rr, arp=rt[r], arp_nbr=len(rlist) - 2,
                                                   is_big_ring='否'))
            else:
                if (len(rlist) - 2) >= 12:
                    rrstr = str(rr) + '-从环' + str(slavenbr)
                    slavenbr -= 1
                    ringtablelist.append(RingTable(region=rr[0:2], ring_name=rrstr, arp=rt[r], arp_nbr=len(rlist) - 2,
                                                   is_big_ring='是'))
                else:
                    rrstr = str(rr) + '-从环' + str(slavenbr)
                    slavenbr -= 1
                    ringtablelist.append(RingTable(region=rr[0:2], ring_name=rrstr, arp=rt[r], arp_nbr=len(rlist) - 2,
                                                   is_big_ring='否'))
    RingTable.objects.bulk_create(ringtablelist)
    print(u"成功生成成环表！")

    # 生成网元表
    neregion = list(set(ConvergeNE.objects.values_list('ring_region', flat=True)))
    nt = NetworkElement.objects.filter(ring_region__in=neregion).values_list('ring_region', 'ne_name')
    rne = DetailResult.objects.exclude(arr_ne='').values_list('arr_ne', flat=True)
    dne = DetailResult.objects.exclude(dr_ne='').values_list('dr_ne', flat=True)
    netablelist = []
    is_ring_or_not = '否'
    is_double_arrive_or_not = '否'
    for n in nt:
        for rn in rne:
            rnlist = str(rn).split('，')
            if n[1] in rnlist:
                is_ring_or_not = '是'
                break
        for dn in dne:
            dnlist = str(dn).split('，')
            if n[1] in dnlist:
                is_double_arrive_or_not = '是'
                break
        netablelist.append(
            NeTable(region=n[0], ne_name=n[1], is_ring=is_ring_or_not, is_double_arrive=is_double_arrive_or_not))
        is_ring_or_not = '否'
        is_double_arrive_or_not = '否'
    NeTable.objects.bulk_create(netablelist)
    print(u"成功生成网元表！")
    return render_to_response('success.html')


@login_required
def download(request):
    netable = NeTable.objects.values_list('region', 'ne_name', 'is_ring', 'is_double_arrive')
    cnetable = CneTable.objects.values_list('region', 'cne_name', 'cne_nenbr', 'is_big_cne')
    ringtable = RingTable.objects.values_list('region', 'ring_name', 'arp', 'arp_nbr', 'is_big_ring')
    longsingletable = LongSingleTable.objects.values_list('region', 'longsinglepath', 'nbr')
    wb_netable = Workbook(write_only=True)
    ws_netable = wb_netable.create_sheet()
    ws_netable.append(['区域', '网元名称', '是否成环', '是否双归'])
    for ne in netable:
        ws_netable.append(ne)
    wb_netable.save('E:\\upload\\netable.xlsx')

    wb_cnetable = Workbook(write_only=True)
    ws_cnetable = wb_cnetable.create_sheet()
    ws_cnetable.append(['区域', '汇聚网元名称', '汇聚网元分值', '是否超大汇聚节点'])
    for cne in cnetable:
        ws_cnetable.append(cne)
    wb_cnetable.save('E:\\upload\\cnetable.xlsx')

    wb_ringtable = Workbook(write_only=True)
    ws_ringtable = wb_ringtable.create_sheet()
    ws_ringtable.append(['区域', '环名称', '成环路经', '成环网元数', '是否超大接入环'])
    for r in ringtable:
        ws_ringtable.append(r)
    wb_ringtable.save('E:\\upload\\ringtable.xlsx')

    wb_longsingletable = Workbook(write_only=True)
    ws_longsingletable = wb_longsingletable.create_sheet()
    ws_longsingletable.append(['区域', '长单链路径', '长单链网元数'])
    for l in longsingletable:
        ws_longsingletable.append(l)
    wb_longsingletable.save('E:\\upload\\longsingletable.xlsx')
    return render_to_response('success.html')
